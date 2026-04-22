#!/usr/bin/env python3
"""
Task 5: Automated Politician Portfolio Tracking & Email Notification System
=========================================================================
Fetches congressional trading data from Quiver Quantitative API,
reconstructs portfolios, detects position changes, generates a styled
3-sheet Excel report, and sends it as an email attachment via SendGrid SMTP
(smtplib, no SendGrid SDK).

Requirements:
    pip install -r requirements.txt

Configuration (env vars or .env file):
    QUIVER_API_TOKEN   — Quiver Quantitative API token
    QUIVER_CSRF_TOKEN  — X-CSRFToken for Quiver API requests
    SENDGRID_API_KEY   — SendGrid API key (SMTP password; username is literal "apikey")
    SENDER_EMAIL       — From address
    RECIPIENT_EMAILS   — Comma-separated recipient list
"""

import requests
import pandas as pd
import smtplib
import json
import os
import sys
import logging
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)


# ============================================================
# TERMINAL STEP PRINTER
# ============================================================
_STEP_COUNTER = [0]

def step(msg: str):
    _STEP_COUNTER[0] += 1
    line = "=" * 60
    print(f"\n{line}")
    print(f"  STEP {_STEP_COUNTER[0]}: {msg}")
    print(f"{line}")

def substep(msg: str):
    print(f"    ✔  {msg}")


# ============================================================
# CONFIGURATION
# ============================================================
QUIVER_API_TOKEN   = os.getenv("QUIVER_API_TOKEN", "")
QUIVER_CSRF_TOKEN  = os.getenv("QUIVER_CSRF_TOKEN", "")
SENDGRID_API_KEY   = os.getenv("SENDGRID_API_KEY", "")
SENDGRID_SMTP_HOST = "smtp.sendgrid.net"
SENDGRID_SMTP_USER = "apikey"  # literal; SendGrid SMTP always uses this username
SENDGRID_SMTP_PORT = 587
SENDER_EMAIL       = os.getenv("SENDER_EMAIL", "")
RECIPIENT_EMAILS   = [r.strip() for r in os.getenv("RECIPIENT_EMAILS", "").split(",") if r.strip()]

POLITICIANS  = ["Nancy Pelosi", "Daniel Meuser"]
SNAPSHOT_DIR = Path("snapshots")
SNAPSHOT_DIR.mkdir(exist_ok=True)

RANGE_MIDPOINTS = {
    1.0: 500, 9.0: 5000, 15.0: 7500, 1001.0: 8000,
    15001.0: 32500, 50001.0: 75000, 100001.0: 175000,
    250001.0: 375000, 500001.0: 750000, 1000001.0: 3000000,
    5000001.0: 15000000
}
CONGRESS_BUYS_LOOKBACK_DAYS = 365

def quiver_headers() -> dict:
    return {
        "accept": "application/json",
        "Authorization": f"Token {QUIVER_API_TOKEN}",
        "X-CSRFToken": QUIVER_CSRF_TOKEN,
    }

POLITICIAN_META = {
    "Nancy Pelosi":  {"party": "Democratic", "chamber": "House of Representatives"},
    "Daniel Meuser": {"party": "Republican", "chamber": "House of Representatives"}
}


# ============================================================
# MODULE 1: FETCH & RECONSTRUCT PORTFOLIO DATA
# ============================================================
def fetch_trades(politician_name: str) -> pd.DataFrame:
    encoded = politician_name.replace(" ", "%20")
    url = f"https://api.quiverquant.com/beta/bulk/congresstrading?representative={encoded}"
    logger.info(f"Fetching trades for {politician_name}")
    r = requests.get(url, headers=quiver_headers(), timeout=30)
    r.raise_for_status()
    df = pd.DataFrame(r.json())
    substep(f"Fetched {len(df)} trade records for {politician_name}")
    return df


def reconstruct_portfolio(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Trade_Size_USD"] = df["Trade_Size_USD"].astype(float)
    df["Traded"]   = pd.to_datetime(df["Traded"], errors="coerce")
    df["midpoint"] = df["Trade_Size_USD"].map(RANGE_MIDPOINTS).fillna(df["Trade_Size_USD"])

    positions = {}
    for _, row in df.sort_values("Traded").iterrows():
        t = row["Ticker"]
        if t not in positions:
            positions[t] = {"value": 0, "last_date": None, "last_action": None, "name": None}
        if positions[t]["name"] is None:
            positions[t]["name"] = row.get("Description", "")
        if row["Transaction"] == "Purchase":
            positions[t]["value"] += row["midpoint"]
        elif row["Transaction"] in ["Sale", "Exchange"]:
            positions[t]["value"] -= row["midpoint"]
        positions[t]["last_date"]   = row["Traded"]
        positions[t]["last_action"] = row["Transaction"]

    holdings = [
        {"Ticker": t, "StockName": d["name"] or "", "EstimatedValue": d["value"],
         "LastTradeDate": d["last_date"], "LastAction": d["last_action"]}
        for t, d in positions.items() if d["value"] > 0
    ]
    result = pd.DataFrame(holdings)
    if len(result):
        total = result["EstimatedValue"].sum()
        result["PortfolioPct"] = (result["EstimatedValue"] / total * 100).round(2)
        result = result.sort_values("EstimatedValue", ascending=False).reset_index(drop=True)
    return result


# ============================================================
# MODULE 2: POSITION CHANGE MONITORING
# ============================================================
def get_snapshot_path(name: str) -> Path:
    return SNAPSHOT_DIR / f"{name.replace(' ', '_').lower()}_snapshot.json"

def load_previous_snapshot(name: str) -> dict:
    p = get_snapshot_path(name)
    return json.load(open(p)) if p.exists() else {}

def save_snapshot(name: str, portfolio: pd.DataFrame):
    snapshot = {
        row["Ticker"]: {
            "value": row["EstimatedValue"], "pct": row["PortfolioPct"],
            "last_date": row["LastTradeDate"].strftime("%Y-%m-%d") if pd.notna(row["LastTradeDate"]) else "N/A",
            "last_action": row["LastAction"]
        }
        for _, row in portfolio.iterrows()
    }
    json.dump(snapshot, open(get_snapshot_path(name), "w"), indent=2)
    substep(f"Snapshot saved → {get_snapshot_path(name)}")

def detect_changes(name: str, portfolio: pd.DataFrame) -> list:
    previous = load_previous_snapshot(name)
    curr, prev = set(portfolio["Ticker"]), set(previous)
    changes = []
    for t in curr - prev:
        row = portfolio[portfolio["Ticker"] == t].iloc[0]
        changes.append({"ticker": t, "type": "NEW POSITION",
                        "old_pct": 0, "new_pct": row["PortfolioPct"], "value": row["EstimatedValue"]})
    for t in prev - curr:
        changes.append({"ticker": t, "type": "POSITION CLOSED",
                        "old_pct": previous[t]["pct"], "new_pct": 0, "value": 0})
    for t in curr & prev:
        row = portfolio[portfolio["Ticker"] == t].iloc[0]
        old, new = previous[t]["pct"], row["PortfolioPct"]
        if abs(new - old) >= 0.5:
            changes.append({"ticker": t, "type": "INCREASED" if new > old else "DECREASED",
                            "old_pct": old, "new_pct": new, "value": row["EstimatedValue"]})
    return changes


# ============================================================
# MODULE 3: CONGRESS BUYS STRATEGY
# ============================================================
def fetch_all_congress_trades() -> pd.DataFrame:
    r = requests.get("https://api.quiverquant.com/beta/live/congresstrading",
                     headers=quiver_headers(), timeout=30)
    r.raise_for_status()
    df = pd.DataFrame(r.json())
    substep(f"Fetched {len(df)} Congress trade records")
    return df

def compute_congress_buys_strategy(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["TransactionDate"] = pd.to_datetime(df["TransactionDate"], errors="coerce")
    df["Amount"]   = pd.to_numeric(df["Amount"], errors="coerce").fillna(0)
    df["midpoint"] = df["Amount"].map(RANGE_MIDPOINTS).fillna(df["Amount"])
    cutoff = df["TransactionDate"].max() - pd.Timedelta(days=CONGRESS_BUYS_LOOKBACK_DAYS)
    buys = df[df["Transaction"].str.contains("Purchase", case=False, na=False)
              & (df["TransactionDate"] >= cutoff)]
    s = buys.groupby("Ticker").agg(
        TotalPurchaseAmount=("midpoint",       "sum"),
        TradeCount=         ("midpoint",       "count"),
        UniqueMembers=      ("Representative", "nunique"),
        LastPurchaseDate=   ("TransactionDate","max"),
        StockName=          ("Description",    "first")
    ).reset_index()
    s = s[s["TotalPurchaseAmount"] > 0].sort_values(
        "TotalPurchaseAmount", ascending=False).reset_index(drop=True)
    s["PortfolioPct"] = (s["TotalPurchaseAmount"] / s["TotalPurchaseAmount"].sum() * 100).round(2)
    return s

def get_congress_buys_snapshot_path() -> Path:
    return SNAPSHOT_DIR / "congress_buys_snapshot.json"

def load_congress_buys_snapshot() -> dict:
    p = get_congress_buys_snapshot_path()
    return json.load(open(p)) if p.exists() else {}

def save_congress_buys_snapshot(strategy: pd.DataFrame):
    snapshot = {
        row["Ticker"]: {
            "pct": row["PortfolioPct"], "amount": row["TotalPurchaseAmount"],
            "trades": int(row["TradeCount"]), "members": int(row["UniqueMembers"]),
            "last_date": pd.to_datetime(row["LastPurchaseDate"]).strftime("%Y-%m-%d")
                if pd.notna(row["LastPurchaseDate"]) else "N/A"
        }
        for _, row in strategy.iterrows()
    }
    json.dump(snapshot, open(get_congress_buys_snapshot_path(), "w"), indent=2)
    substep(f"Congress Buys snapshot saved → {get_congress_buys_snapshot_path()}")

def detect_congress_buys_changes(strategy: pd.DataFrame) -> list:
    previous = load_congress_buys_snapshot()
    curr, prev = set(strategy["Ticker"]), set(previous)
    changes = []
    for t in curr - prev:
        row = strategy[strategy["Ticker"] == t].iloc[0]
        changes.append({"ticker": t, "type": "NEW IN STRATEGY",
                        "old_pct": 0.0, "new_pct": row["PortfolioPct"]})
    for t in prev - curr:
        changes.append({"ticker": t, "type": "DROPPED FROM STRATEGY",
                        "old_pct": previous[t]["pct"], "new_pct": 0.0})
    for t in curr & prev:
        row = strategy[strategy["Ticker"] == t].iloc[0]
        old, new = previous[t]["pct"], row["PortfolioPct"]
        if abs(new - old) >= 0.5:
            changes.append({"ticker": t,
                            "type": "WEIGHT INCREASED" if new > old else "WEIGHT DECREASED",
                            "old_pct": old, "new_pct": new})
    return changes


# ============================================================
# MODULE 4: EXCEL REPORT GENERATION (3 sheets)
# ============================================================
DARK_NAVY   = "1A1A2E"
MID_NAVY    = "16213E"
ACCENT_BLUE = "0F3460"
CONGRESS_BG = "0F4C75"
BUY_GREEN   = "27AE60"
SELL_RED    = "E74C3C"
LIGHT_GRAY  = "F8F9FA"
ALT_ROW     = "EEF2F7"
HEADER_GOLD = "E8C96A"

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(size=10, bold=False, color="333333"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_title(ws, title, subtitle, generated_at, ncols):
    ws.row_dimensions[1].height = 36
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = title
    c.font  = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    c.fill  = _fill(DARK_NAVY)
    c.alignment = _center()

    ws.row_dimensions[2].height = 22
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c = ws["A2"]
    c.value = f"{subtitle}   |   Generated: {generated_at}"
    c.font  = Font(name="Arial", size=10, color="AAAAAA")
    c.fill  = _fill(MID_NAVY)
    c.alignment = _center()

    ws.row_dimensions[3].height = 6
    ws.merge_cells(f"A3:{get_column_letter(ncols)}3")
    ws["A3"].fill = _fill("F0F4F8")


def _write_info_bar(ws, row, text, ncols, bg="E8F0FE"):
    ws.row_dimensions[row].height = 20
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font  = Font(name="Arial", size=10, bold=True, color=DARK_NAVY)
    c.fill  = _fill(bg)
    c.alignment = _left()
    c.border = _thin_border()


def _write_col_headers(ws, row, headers, bg=DARK_NAVY):
    ws.row_dimensions[row].height = 30
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill      = _fill(bg)
        c.alignment = _center()
        c.border    = _thin_border()


def _write_totals_row(ws, row, ncols, value_col, bg=DARK_NAVY):
    ws.row_dimensions[row].height = 22
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = _fill(bg); c.border = _thin_border()
    ws.cell(row=row, column=1).value     = "TOTAL"
    ws.cell(row=row, column=1).font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).alignment = _center()
    c = ws.cell(row=row, column=value_col,
                value=f"=SUM({get_column_letter(value_col)}6:{get_column_letter(value_col)}{row-1})")
    c.number_format = '$#,##0'
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.alignment = _center()
    c2 = ws.cell(row=row, column=value_col + 1, value=1.0)
    c2.number_format = '0.00%'
    c2.font = Font(name="Arial", size=10, bold=True, color=HEADER_GOLD)
    c2.alignment = _center()


def _write_disclaimer(ws, row, text, ncols=8):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font  = Font(name="Arial", size=8, italic=True, color="888888")
    c.alignment = _left()


def build_politician_sheet(wb: Workbook, sheet_name: str, politician_name: str,
                            party: str, chamber: str, portfolio: pd.DataFrame):
    ws  = wb.create_sheet(title=sheet_name)
    now = datetime.now().strftime("%Y-%m-%d %H:%M UTC")

    for i, w in enumerate([8, 38, 18, 16, 16, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _write_title(ws, f"{politician_name} — Stock Portfolio",
                 f"{party}  |  {chamber}", now, ncols=6)

    total_val = portfolio["EstimatedValue"].sum()
    _write_info_bar(ws, 4,
        f"Total Estimated Portfolio Value:   ${total_val:,.0f}   |   "
        f"Open Positions: {len(portfolio)}", ncols=6)

    _write_col_headers(ws, 5,
        ["Ticker", "Stock Name", "Est. Value ($)", "% of Portfolio",
         "Last Action", "Last Trade Date"])

    for i, (_, row) in enumerate(portfolio.iterrows()):
        r  = i + 6
        bg = LIGHT_GRAY if i % 2 == 0 else ALT_ROW
        ws.row_dimensions[r].height = 18

        c = ws.cell(row=r, column=1, value=row["Ticker"])
        c.font = Font(name="Arial", size=10, bold=True, color=ACCENT_BLUE)
        c.fill = _fill(bg); c.alignment = _center(); c.border = _thin_border()

        c = ws.cell(row=r, column=2, value=row["StockName"])
        c.font = _font(); c.fill = _fill(bg); c.alignment = _left(); c.border = _thin_border()

        c = ws.cell(row=r, column=3, value=row["EstimatedValue"])
        c.number_format = '$#,##0'
        c.font = _font(bold=True); c.fill = _fill(bg); c.alignment = _center(); c.border = _thin_border()

        c = ws.cell(row=r, column=4, value=row["PortfolioPct"] / 100)
        c.number_format = '0.00%'
        c.font = _font(); c.fill = _fill(bg); c.alignment = _center(); c.border = _thin_border()

        action = row["LastAction"]
        c = ws.cell(row=r, column=5, value=action)
        c.font = Font(name="Arial", size=10, bold=True,
                      color=BUY_GREEN if action == "Purchase" else SELL_RED)
        c.fill = _fill(bg); c.alignment = _center(); c.border = _thin_border()

        date_str = row["LastTradeDate"].strftime("%Y-%m-%d") \
            if pd.notna(row["LastTradeDate"]) else "N/A"
        c = ws.cell(row=r, column=6, value=date_str)
        c.font = _font(); c.fill = _fill(bg); c.alignment = _center(); c.border = _thin_border()

    total_row = len(portfolio) + 6
    _write_totals_row(ws, total_row, ncols=6, value_col=3)
    ws.freeze_panes = "A6"
    _write_disclaimer(ws, total_row + 2,
        "⚠  Data sourced from STOCK Act congressional financial disclosure filings via "
        "Quiver Quantitative API. Values are estimates based on disclosed dollar ranges. "
        "Not investment advice.", ncols=6)


def build_congress_sheet(wb: Workbook, strategy: pd.DataFrame):
    ws  = wb.create_sheet(title="Congress Buys Strategy")
    now = datetime.now().strftime("%Y-%m-%d %H:%M UTC")

    for i, w in enumerate([8, 38, 22, 16, 12, 16, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _write_title(ws,
        "Congress Buys Strategy — All Holdings",
        "All U.S. Congress Members  |  Weighted by Purchase Size  |  Rolling 12-Month Window",
        now, ncols=7)

    total_vol = strategy["TotalPurchaseAmount"].sum()
    _write_info_bar(ws, 4,
        f"Total Purchase Volume Tracked: ${total_vol:,.0f} (est.)   |   "
        f"Holdings: {len(strategy)}   |   "
        f"Strategy CAGR (Quiver backtest since 2020-04-01): 33.01%",
        ncols=7, bg="E0F0FF")

    _write_col_headers(ws, 5,
        ["Ticker", "Stock Name", "Est. Purchase Vol. ($)", "% of Strategy",
         "# Trades", "# Members", "Last Buy Date"],
        bg=CONGRESS_BG)

    for i, (_, row) in enumerate(strategy.iterrows()):
        r  = i + 6
        bg = LIGHT_GRAY if i % 2 == 0 else ALT_ROW
        ws.row_dimensions[r].height = 18

        c = ws.cell(row=r, column=1, value=row["Ticker"])
        c.font = Font(name="Arial", size=10, bold=True, color=CONGRESS_BG)
        c.fill = _fill(bg); c.alignment = _center(); c.border = _thin_border()

        c = ws.cell(row=r, column=2, value=row.get("StockName", ""))
        c.font = _font(); c.fill = _fill(bg); c.alignment = _left(); c.border = _thin_border()

        c = ws.cell(row=r, column=3, value=row["TotalPurchaseAmount"])
        c.number_format = '$#,##0'
        c.font = _font(bold=True); c.fill = _fill(bg); c.alignment = _center(); c.border = _thin_border()

        c = ws.cell(row=r, column=4, value=row["PortfolioPct"] / 100)
        c.number_format = '0.00%'
        c.font = _font(); c.fill = _fill(bg); c.alignment = _center(); c.border = _thin_border()

        c = ws.cell(row=r, column=5, value=int(row["TradeCount"]))
        c.font = _font(); c.fill = _fill(bg); c.alignment = _center(); c.border = _thin_border()

        c = ws.cell(row=r, column=6, value=int(row["UniqueMembers"]))
        c.font = _font(); c.fill = _fill(bg); c.alignment = _center(); c.border = _thin_border()

        date_str = pd.to_datetime(row["LastPurchaseDate"]).strftime("%Y-%m-%d") \
            if pd.notna(row["LastPurchaseDate"]) else "N/A"
        c = ws.cell(row=r, column=7, value=date_str)
        c.font = _font(); c.fill = _fill(bg); c.alignment = _center(); c.border = _thin_border()

    total_row = len(strategy) + 6
    _write_totals_row(ws, total_row, ncols=7, value_col=3, bg=CONGRESS_BG)
    ws.freeze_panes = "A6"
    _write_disclaimer(ws, total_row + 2,
        "⚠  Data sourced from STOCK Act filings via Quiver Quantitative API. "
        "Quiver-reported CAGR of 33.01% based on backtest from 2020-04-01. "
        "Past performance does not guarantee future results. Not investment advice.", ncols=7)


def generate_excel_report(politician_portfolios: dict,
                           congress_strategy: pd.DataFrame,
                           excel_path: Path) -> bool:
    """Build the 3-sheet Excel workbook and save it to excel_path."""
    try:
        wb = Workbook()
        wb.remove(wb.active)  # remove default blank sheet

        for name, portfolio in politician_portfolios.items():
            meta = POLITICIAN_META.get(name, {"party": "Unknown", "chamber": "Unknown"})
            build_politician_sheet(wb, name, name, meta["party"], meta["chamber"], portfolio)
            substep(f"Sheet '{name}' written — {len(portfolio)} positions")

        build_congress_sheet(wb, congress_strategy)
        substep(f"Sheet 'Congress Buys Strategy' written — {len(congress_strategy)} holdings")

        wb.save(str(excel_path))
        size_kb = excel_path.stat().st_size // 1024
        print(f"    ✅ Excel report saved → {excel_path.resolve()}  ({size_kb} KB)")
        logger.info(f"Excel report saved at {excel_path}")
        return True
    except Exception as e:
        print(f"    ❌ Excel generation failed: {e}")
        logger.error(f"Excel generation error: {e}", exc_info=True)
        return False


# ============================================================
# MODULE 5: EMAIL DELIVERY VIA SENDGRID SMTP
# ============================================================
def send_email_sendgrid(subject_prefix: str, excel_path: Path):
    """Send a short notification email with the Excel report attached."""
    if not SENDGRID_API_KEY:
        print("\n  ⚠  SendGrid API key missing — skipping email send.")
        return

    now_str    = datetime.now().strftime("%Y-%m-%d at %H:%M UTC")
    subject    = f"{subject_prefix} — Congress Portfolio Report — {datetime.now().strftime('%Y-%m-%d')}"
    recipients = [r.strip() for r in RECIPIENT_EMAILS if r.strip()]
    if not recipients:
        print("\n  ⚠  No recipient emails configured — skipping send.")
        return

    text_body = (
        f"Congressional Portfolio Report\n\n"
        f"This report was generated on {now_str}.\n\n"
        f"Please open the attached Excel file to view the full report.\n"
        f"It contains three sheets:\n"
        f"  - Nancy Pelosi — individual stock portfolio\n"
        f"  - Daniel Meuser — individual stock portfolio\n"
        f"  - Congress Buys Strategy — all holdings weighted by purchase volume\n\n"
        f"— Automated Politician Portfolio Monitoring System"
    )
    html_body = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="font-family:'Segoe UI',Arial,sans-serif;background:#f5f5f5;margin:0;padding:30px">
  <div style="max-width:600px;margin:0 auto;background:#fff;border-radius:8px;
              overflow:hidden;box-shadow:0 2px 10px rgba(0,0,0,0.1)">
    <div style="background:linear-gradient(135deg,#1a1a2e 0%,#16213e 100%);
                color:#fff;padding:30px">
      <h1 style="margin:0;font-size:22px">{subject_prefix}</h1>
      <p style="margin:8px 0 0;opacity:0.8;font-size:13px">
        Automated Politician Portfolio Monitoring System
      </p>
    </div>
    <div style="padding:30px">
      <p style="font-size:16px;color:#333;margin-top:0">
        This report was generated on <strong>{now_str}</strong>.
      </p>
      <p style="font-size:15px;color:#555">
        Please open the <strong>attached Excel file</strong> to view the full report.
        It contains three sheets:
      </p>
      <ul style="color:#555;font-size:15px;line-height:2">
        <li><strong>Nancy Pelosi</strong> — individual stock portfolio</li>
        <li><strong>Daniel Meuser</strong> — individual stock portfolio</li>
        <li><strong>Congress Buys Strategy</strong> — all holdings weighted by purchase volume</li>
      </ul>
    </div>
    <div style="padding:15px 30px;background:#1a1a2e;color:#888;
                font-size:11px;text-align:center">
      Automated Portfolio Tracking System | Powered by Quiver Quantitative API
      | Delivered via SendGrid
    </div>
  </div>
</body></html>"""

    print(f"    From       : {SENDER_EMAIL}")
    print(f"    To         : {', '.join(recipients)}")
    print(f"    Subject    : {subject}")
    print(f"    Attachment : {excel_path.name}  ({excel_path.stat().st_size // 1024} KB)")

    excel_bytes = excel_path.read_bytes()

    for recipient in recipients:
        print(f"\n    → Connecting to SendGrid SMTP for: {recipient}")
        msg            = MIMEMultipart("mixed")
        msg["Subject"] = subject
        msg["From"]    = SENDER_EMAIL
        msg["To"]      = recipient

        body = MIMEMultipart("alternative")
        body.attach(MIMEText(text_body, "plain"))
        body.attach(MIMEText(html_body, "html"))
        msg.attach(body)

        part = MIMEApplication(excel_bytes,
               _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.add_header("Content-Disposition", "attachment", filename=excel_path.name)
        msg.attach(part)

        try:
            with smtplib.SMTP(SENDGRID_SMTP_HOST, SENDGRID_SMTP_PORT) as server:
                server.ehlo(); server.starttls(); server.ehlo()
                server.login(SENDGRID_SMTP_USER, SENDGRID_API_KEY)
                server.sendmail(SENDER_EMAIL, recipient, msg.as_string())
            print(f"    ✅ Email successfully sent to: {recipient}")
            logger.info(f"Email sent to {recipient}")
        except smtplib.SMTPAuthenticationError as e:
            print(f"    ❌ Authentication failed: {e}")
            logger.error(f"SendGrid auth failed: {e}")
        except smtplib.SMTPException as e:
            print(f"    ❌ SMTP error: {e}")
            logger.error(f"SMTP error: {e}")
        except Exception as e:
            print(f"    ❌ Unexpected error: {e}")
            logger.error(f"Email error: {e}", exc_info=True)


# ============================================================
# MAIN ORCHESTRATION
# ============================================================
def run_portfolio_check():
    _STEP_COUNTER[0] = 0
    print("\n" + "█" * 60)
    print("  POLITICIAN PORTFOLIO TRACKING SYSTEM — STARTING RUN")
    print("  " + datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC"))
    print("█" * 60)

    politician_portfolios = {}
    any_changes = False

    # ── Individual politician portfolios ─────────────────────────────────────
    for politician in POLITICIANS:
        step(f"PROCESSING POLITICIAN: {politician.upper()}")
        try:
            print("    Fetching trade history from Quiver API...")
            trades    = fetch_trades(politician)
            print("    Reconstructing current portfolio...")
            portfolio = reconstruct_portfolio(trades)

            if portfolio.empty:
                print(f"    ⚠  No open positions found for {politician} — skipping.")
                continue

            substep(f"Portfolio reconstructed — {len(portfolio)} open positions")

            print("    Detecting position changes vs. last snapshot...")
            changes = detect_changes(politician, portfolio)
            if changes:
                any_changes = True
                print(f"    ⚠  {len(changes)} change(s) detected:")
                for c in changes:
                    print(f"         {c['ticker']:6s}  {c['type']}  "
                          f"{c['old_pct']:.2f}% → {c['new_pct']:.2f}%")
            else:
                substep(f"No portfolio changes detected for {politician}")

            politician_portfolios[politician] = portfolio
            print("    Saving snapshot...")
            save_snapshot(politician, portfolio)

        except requests.exceptions.RequestException as e:
            print(f"    ❌ API error for {politician}: {e}")
            logger.error(f"API error for {politician}: {e}")
        except Exception as e:
            print(f"    ❌ Unexpected error for {politician}: {e}")
            logger.error(f"Error processing {politician}: {e}", exc_info=True)

    # ── Congress Buys Strategy ────────────────────────────────────────────────
    step("PROCESSING CONGRESS BUYS STRATEGY")
    cb_strategy = None
    try:
        print("    Fetching all Congress trade records from Quiver API...")
        cb_trades = fetch_all_congress_trades()
        print("    Computing Congress Buys portfolio (12-month rolling window)...")
        cb_strategy = compute_congress_buys_strategy(cb_trades)
        substep(f"Congress Buys strategy computed — {len(cb_strategy)} holdings")

        print("    Detecting Congress Buys strategy changes...")
        cb_changes = detect_congress_buys_changes(cb_strategy)
        if cb_changes:
            any_changes = True
            print(f"    ⚠  {len(cb_changes)} strategy change(s) detected:")
            for c in cb_changes:
                print(f"         {c['ticker']:6s}  {c['type']}  "
                      f"{c['old_pct']:.2f}% → {c['new_pct']:.2f}%")
        else:
            substep("No strategy changes detected")

        print("    Saving Congress Buys snapshot...")
        save_congress_buys_snapshot(cb_strategy)

    except Exception as e:
        print(f"    ❌ Error processing Congress Buys Strategy: {e}")
        logger.error(f"Error processing Congress Buys Strategy: {e}", exc_info=True)

    # ── Generate Excel + send email ───────────────────────────────────────────
    if politician_portfolios and cb_strategy is not None:
        subject_prefix = "ALERT: Position Change" if any_changes else "Portfolio Summary"

        step("GENERATING EXCEL REPORT (3 sheets)")
        excel_path = SNAPSHOT_DIR / "congress_portfolio_report.xlsx"
        excel_ok   = generate_excel_report(politician_portfolios, cb_strategy, excel_path)

        if excel_ok:
            step("SENDING EMAIL WITH EXCEL ATTACHMENT")
            if any_changes:
                print("    Position/strategy changes detected — sending ALERT email.")
            else:
                print("    No changes — sending full PORTFOLIO SUMMARY email.")
            send_email_sendgrid(subject_prefix, excel_path)
        else:
            print("    ⚠  Excel generation failed — email not sent.")

    print("\n" + "█" * 60)
    print("  PORTFOLIO CHECK COMPLETE")
    print("  " + datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC"))
    print("█" * 60 + "\n")
    logger.info("Portfolio check complete.")
    sys.exit(0)


def run_weekly_summary():
    _STEP_COUNTER[0] = 0
    print("\n" + "█" * 60)
    print("  WEEKLY PORTFOLIO SUMMARY RUN — STARTING")
    print("  " + datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC"))
    print("█" * 60)

    politician_portfolios = {}

    for politician in POLITICIANS:
        step(f"PROCESSING POLITICIAN: {politician.upper()}")
        try:
            print("    Fetching trade history...")
            trades    = fetch_trades(politician)
            print("    Reconstructing portfolio...")
            portfolio = reconstruct_portfolio(trades)
            substep(f"{len(portfolio)} open positions")
            politician_portfolios[politician] = portfolio
            print("    Saving snapshot...")
            save_snapshot(politician, portfolio)
        except Exception as e:
            print(f"    ❌ Error for {politician}: {e}")
            logger.error(f"Error for {politician}: {e}", exc_info=True)

    step("PROCESSING CONGRESS BUYS STRATEGY")
    cb_strategy = None
    try:
        print("    Fetching all Congress trade records...")
        cb_trades = fetch_all_congress_trades()
        print("    Computing strategy...")
        cb_strategy = compute_congress_buys_strategy(cb_trades)
        substep(f"{len(cb_strategy)} holdings")
        save_congress_buys_snapshot(cb_strategy)
    except Exception as e:
        print(f"    ❌ Error processing Congress Buys Strategy: {e}")
        logger.error(f"Error processing Congress Buys Strategy: {e}", exc_info=True)

    if politician_portfolios and cb_strategy is not None:
        step("GENERATING EXCEL REPORT (3 sheets)")
        excel_path = SNAPSHOT_DIR / "congress_portfolio_report.xlsx"
        excel_ok   = generate_excel_report(politician_portfolios, cb_strategy, excel_path)

        if excel_ok:
            step("SENDING EMAIL WITH EXCEL ATTACHMENT")
            send_email_sendgrid("Weekly Portfolio Summary", excel_path)
        else:
            print("    ⚠  Excel generation failed — email not sent.")

    print("\n" + "█" * 60)
    print("  WEEKLY SUMMARY COMPLETE")
    print("  " + datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC"))
    print("█" * 60 + "\n")
    sys.exit(0)


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--summary":
        run_weekly_summary()
    else:
        run_portfolio_check()